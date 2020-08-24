from transformers.tokenization_auto import AutoTokenizer
from transformers.modeling_auto import AutoModelForSequenceClassification
from transformers import Trainer, TrainingArguments, AutoConfig
from itertools import islice
from openpyxl import load_workbook
import random
from torch.utils.data import Dataset
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
import torch

fn = "wellness.xlsx"
wb = load_workbook(filename=fn)
ws = wb[wb.sheetnames[0]]

labels = {}
label_map = {}
answers = {}
for row in islice(ws.iter_rows(), 1, None):
    label = row[0].value.strip()#.split('/')[0]
    label_map.setdefault(label, len(label_map))
    utter = row[1].value.strip()
    labels.setdefault(label_map[label], []).append(utter)
    if row[2].value is not None:
        answers.setdefault(label_map[label], []).append(row[2].value.strip())

class WellnessDataset(Dataset):
    def __init__(self, data):
        self.data = data

    def __len__(self):
        return len(self.data)

    def __getitem__(self, index):
        return self.data[index]

def gen_datasets(split=0.1):
    train_dataset, test_dataset = [], []
    train_label, test_label = [], []
    for k in labels:
        random.shuffle(labels[k])
        assert len(labels[k]) > 1
        ix = max(1, int(len(labels[k])*split))
        for i in range(len(labels[k])):
            if i < ix:
                test_dataset.append(labels[k][i])
                test_label.append(k)
            else:
                train_dataset.append(labels[k][i])
                train_label.append(k)
    return train_dataset, train_label, test_dataset, test_label

train_dataset, train_label, test_dataset, test_label = gen_datasets()

model_name_or_path = "monologg/koelectra-base-discriminator"
#model_name_or_path = "monologg/kobert"

tokenizer = AutoTokenizer.from_pretrained(model_name_or_path)

import numpy as np
train_dataset = tokenizer(train_dataset, return_tensors='pt', padding='max_length', truncation=True)
test_dataset = tokenizer(test_dataset, return_tensors='pt', padding='max_length', truncation=True)

train_dataset['label'] = train_label
test_dataset['label'] =  test_label

keys = list(train_dataset.keys())
train_dataset = [dict(zip(keys, v)) for v in zip(*train_dataset.values())]
test_dataset = [dict(zip(keys, v)) for v in zip(*test_dataset.values())]

train_dataset = WellnessDataset(train_dataset)
test_dataset = WellnessDataset(test_dataset)

print("train_dataset:", len(train_dataset))
print("test dataset size:", len(test_dataset))

config = AutoConfig.from_pretrained(model_name_or_path, num_labels=len(label_map))

model = AutoModelForSequenceClassification.from_pretrained(model_name_or_path, config=config)
#for param in model.base_model.parameters():
#    param.requires_grad = False
#model.train()
print(model)

#no_decay = ['bias', 'LayerNorm.weight']
#optimizer_grouped_parameters = [
#        {'params': [p for n, p in model.named_parameters() if not any(nd in n for nd in no_decay)], 'weight_decay':
#         0.01},
#        {'params': [p for n, p in model.named_parameters() if any(nd in n for nd in no_decay)], 'weight_decay': 0.0}
#]
#optimizer = AdamW(optimizer_grouped_parameters, lr=1e-5)



def compute_metrics(pred):
    labels = pred.label_ids
    preds = pred.predictions.argmax(-1)
    precision, recall, f1, _ = precision_recall_fscore_support(labels, preds, average='micro')
    acc = accuracy_score(labels, preds)

    return {
        'accuracy': acc,
        'f1': f1,
        'precision': precision,
        'recall': recall
    }

training_args = TrainingArguments(
    output_dir='./results',          # output directory
    num_train_epochs=60,             # total # of training epochs
    per_device_train_batch_size=16,  # batch size per device during training
    per_device_eval_batch_size=128,  # batch size for evaluation
    warmup_steps=500,                # number of warmup steps for learning rate scheduler
    weight_decay=0.01,               # strength of weight decay
    logging_dir='./logs',            # directory for storing logs
    evaluate_during_training=True,
    eval_steps=100,
    logging_steps=100,
    do_eval=True,
    do_train=True,
)

trainer = Trainer(
    model=model,                         # the instantiated 🤗 Transformers model to be trained
    args=training_args,                  # training arguments, defined above
    train_dataset=train_dataset,         # training dataset
    eval_dataset=test_dataset,           # evaluation dataset
    compute_metrics=compute_metrics,
)


trainer.train()
print(trainer.evaluate())

torch.save(model, "wellness_model.pt")


