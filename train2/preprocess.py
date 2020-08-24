from transformers.tokenization_auto import AutoTokenizer
from transformers.modeling_auto import AutoModelForSequenceClassification
from transformers import Trainer, TrainingArguments, AutoConfig
from itertools import islice
from openpyxl import load_workbook
import random
from torch.utils.data import Dataset
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
import torch

fn = "wellness.xlsx"
wb = load_workbook(filename=fn)
ws = wb[wb.sheetnames[0]]

labels = {}
label_map = {}
answers = {}
for row in islice(ws.iter_rows(), 1, None):
    label = row[0].value.strip()#.split('/')[0]
    label_map.setdefault(label, len(label_map))
    utter = row[1].value.strip()
    labels.setdefault(label_map[label], []).append(utter)
    if row[2].value is not None:
        answers.setdefault(label_map[label], []).append(row[2].value.strip())
#print(answers)
label_map = {v: k for k, v in label_map.items()}
torch.save(label_map, "label_map.pt")
torch.save(answers, "answers.pt")
