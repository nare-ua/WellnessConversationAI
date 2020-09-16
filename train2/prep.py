import torch
from openpyxl import load_workbook
from itertools import islice
import collections
from pathlib import Path

def load_data(datasetdir=Path('/mnt/data/datasets/wellness'), regen=False):
    p = datasetdir/'preprocessed.pt'
    if regen and p.exists(): p.unlink()
    if not p.exists():
        utters, label_map, answers = parse_data()
        data = {
            'utters': utters,
            'label_map': label_map,
            'answers': answers
        }
        torch.save(data, p)
        return utters, label_map, answers

    data = torch.load(p)
    return data['utters'], data['label_map'], data['answers']


def parse_data(datasetdir=Path('/mnt/data/datasets/wellness')):
    fn = datasetdir/"wellness.xlsx"
    wb = load_workbook(filename=fn)
    ws = wb[wb.sheetnames[0]]

    utters = collections.defaultdict(list)
    answers = collections.defaultdict(list)
    label_map = {}

    for row in islice(ws.iter_rows(), 1, None):
        label = row[0].value.strip()
        utter = row[1].value.strip()

        label_map.setdefault(label, len(label_map))
        utters[label_map[label]].append(utter)

        if row[2].value is not None:
            answers[label_map[label]].append(row[2].value.strip())
    return utters, label_map, answers

if __name__ == "__main__":
    load_data(regen=True)
